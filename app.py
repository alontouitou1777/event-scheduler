import math
from datetime import datetime
from urllib.parse import quote
from io import BytesIO

import streamlit as st
import pandas as pd
import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill


# ======================
# DB (SQLite)
# ======================
DB_PATH = "schedule.db"

def get_conn():
    return sqlite3.connect(DB_PATH, check_same_thread=False)

def init_db():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            phone TEXT NOT NULL,
            role TEXT NOT NULL,         -- waiter / bartender
            days TEXT NOT NULL,         -- "ב,ד,ה"
            rank INTEGER NOT NULL
        )
    """)
    conn.commit()
    conn.close()

def db_has_employees() -> bool:
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM employees")
    n = cur.fetchone()[0]
    conn.close()
    return n > 0

def seed_default_employees():
    defaults = build_default_employees()
    save_employees_to_db(defaults)

def load_employees_from_db() -> list[dict]:
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id, name, phone, role, days, rank FROM employees ORDER BY role, name")
    rows = cur.fetchall()
    conn.close()

    employees = []
    for (eid, name, phone, role, days, rank) in rows:
        days_list = [d.strip() for d in (days or "").split(",") if d.strip()]
        employees.append({
            "id": int(eid),
            "name": name,
            "phone": phone,
            "role": role,
            "days": days_list,
            "rank": int(rank),
        })
    return employees

def save_employees_to_db(employees: list[dict]) -> None:
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM employees")

    for e in employees:
        name = (e.get("name") or "").strip()
        phone = (e.get("phone") or "").strip()
        role = (e.get("role") or "").strip()
        days = e.get("days") or []
        if isinstance(days, list):
            days_str = ",".join([d.strip() for d in days if d.strip()])
        else:
            days_str = str(days)

        rank = int(e.get("rank") or 1)

        if not name:
            continue
        if role not in ("waiter", "bartender"):
            continue

        cur.execute(
            "INSERT INTO employees(name, phone, role, days, rank) VALUES(?,?,?,?,?)",
            (name, phone, role, days_str, rank)
        )

    conn.commit()
    conn.close()


# ======================
# זמן
# ======================
def time_to_minutes(t: str) -> int:
    h, m = t.split(":")
    return int(h) * 60 + int(m)

def minutes_to_time(x: int) -> str:
    x = x % (24 * 60)
    h = x // 60
    m = x % 60
    return f"{h:02d}:{m:02d}"

def overlaps(a_start: int, a_end: int, b_start: int, b_end: int) -> bool:
    return a_start < b_end and b_start < a_end


# ======================
# חוקים
# ======================
def calc_staff(guests: int) -> tuple[int, int]:
    waiters = max(3, math.ceil(guests / 20))
    bartenders = max(1, math.ceil(guests / 60))
    return waiters, bartenders

def difficulty_from_guests(guests: int) -> int:
    if guests <= 120:
        return 2
    if guests <= 250:
        return 3
    if guests <= 400:
        return 4
    return 5


# ======================
# עובדים (ברירת מחדל)
# ======================
def build_default_employees():
    return [
        {"name": "רועי",   "phone": "972500000001", "role": "waiter",    "days": ["ד", "ה", "ו"],      "rank": 1},
        {"name": "עדן",   "phone": "972500000002", "role": "waiter",    "days": ["ב", "ד", "ה", "ו"], "rank": 2},
        {"name": "שי",    "phone": "972500000003", "role": "waiter",    "days": ["ה", "ו", "ש"],      "rank": 2},
        {"name": "נועם",  "phone": "972500000004", "role": "waiter",    "days": ["ג", "ד", "ה"],      "rank": 3},
        {"name": "גל",    "phone": "972500000005", "role": "waiter",    "days": ["ה", "ו", "ש"],      "rank": 3},
        {"name": "אור",   "phone": "972500000006", "role": "waiter",    "days": ["ו", "ש"],           "rank": 4},
        {"name": "אלי",   "phone": "972500000007", "role": "waiter",    "days": ["ב", "ה", "ו"],      "rank": 4},
        {"name": "מיכאל", "phone": "972500000008", "role": "waiter",    "days": ["ה", "ו"],           "rank": 5},

        {"name": "דני",  "phone": "972500000009", "role": "bartender", "days": ["ד", "ה", "ו", "ש"], "rank": 1},
        {"name": "מאור", "phone": "972500000010", "role": "bartender", "days": ["ב", "ד", "ה", "ו"], "rank": 2},
        {"name": "אבי",  "phone": "972500000011", "role": "bartender", "days": ["ה", "ו", "ש"],      "rank": 3},
        {"name": "ברק",  "phone": "972500000012", "role": "bartender", "days": ["ה", "ו"],           "rank": 4},
    ]


# ======================
# אירוע
# ======================
def build_event(day: str, hall: str, guests: int, start_time: str, end_time: str) -> dict:
    waiters, bartenders = calc_staff(guests)

    start_min = time_to_minutes(start_time)
    end_min = time_to_minutes(end_time)
    if end_min <= start_min:
        end_min += 24 * 60  # עבר חצות

    arrival_min = start_min - 120
    difficulty = difficulty_from_guests(guests)

    return {
        "day": day,
        "hall": hall,
        "guests": guests,
        "difficulty": difficulty,

        "start_time": start_time,
        "end_time": end_time,
        "arrival_time": minutes_to_time(arrival_min),

        "arrival_min": arrival_min,
        "start_min": start_min,
        "end_min": end_min,

        "waiters": waiters,
        "bartenders": bartenders,

        "waiter_names": [],
        "bartender_names": [],
        "missing_waiters": 0,
        "missing_bartenders": 0,
    }


# ======================
# שיבוץ
# ======================
def can_assign(name: str, event: dict, schedule: dict) -> bool:
    for d, s, e in schedule.get(name, []):
        if d == event["day"] and overlaps(s, e, event["arrival_min"], event["end_min"]):
            return False
    return True

def assign_event(event: dict, employees: list[dict], assigned: dict, schedule: dict) -> None:
    day = event["day"]
    difficulty = event["difficulty"]

    available = [e for e in employees if day in e["days"]]
    waiters = [e for e in available if e["role"] == "waiter"]
    bartenders = [e for e in available if e["role"] == "bartender"]

    # קושי גבוה -> rank קטן קודם; קושי נמוך -> rank גבוה קודם
    if difficulty >= 4:
        key = lambda e: (assigned.get(e["name"], 0), e["rank"])
    else:
        key = lambda e: (assigned.get(e["name"], 0), -e["rank"])

    waiters.sort(key=key)
    bartenders.sort(key=key)

    chosen_w, chosen_b = [], []

    for e in waiters:
        if len(chosen_w) == event["waiters"]:
            break
        if can_assign(e["name"], event, schedule):
            chosen_w.append(e)

    for e in bartenders:
        if len(chosen_b) == event["bartenders"]:
            break
        if can_assign(e["name"], event, schedule):
            chosen_b.append(e)

    for e in chosen_w + chosen_b:
        assigned[e["name"]] = assigned.get(e["name"], 0) + 1
        schedule.setdefault(e["name"], []).append((day, event["arrival_min"], event["end_min"]))

    event["waiter_names"] = [e["name"] for e in chosen_w]
    event["bartender_names"] = [e["name"] for e in chosen_b]
    event["missing_waiters"] = max(0, event["waiters"] - len(chosen_w))
    event["missing_bartenders"] = max(0, event["bartenders"] - len(chosen_b))


# ======================
# אקסל למנהל (Bytes)
# ======================
def build_manager_excel_bytes(by_day: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "סידור שבועי"
    ws.sheet_view.rightToLeft = True

    ltr = "\u200e"
    days = ["א", "ב", "ג", "ד", "ה", "ו", "ש"]

    YELLOW = PatternFill("solid", fgColor="FFD966")
    header_font = Font(bold=True, size=14)
    normal_font = Font(size=12)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right_top = Alignment(horizontal="right", vertical="top", wrap_text=True)

    thick = Side(style="thick", color="000000")
    border = Border(left=thick, right=thick, top=thick, bottom=thick)

    ws.merge_cells("A1:G1")
    ws["A1"] = "סידור שבועי"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = center

    for i in range(len(days)):
        ws.column_dimensions[chr(ord("A") + i)].width = 32
    ws.row_dimensions[2].height = 26
    ws.row_dimensions[3].height = 420

    for i, day in enumerate(days):
        col = chr(ord("A") + i)

        day_cell = ws[f"{col}2"]
        day_cell.value = f"יום {day}"
        day_cell.font = header_font
        day_cell.fill = YELLOW
        day_cell.alignment = center
        day_cell.border = border

        lines = []
        events = sorted(by_day.get(day, []), key=lambda e: e["start_min"])

        if not events:
            lines.append("—")
        else:
            for ev in events:
                lines.append(f"אולם: {ltr}{ev['hall']}{ltr} | מוזמנים: {ltr}{ev['guests']}{ltr}")
                lines.append(f"הגעה: {ev['arrival_time']} | שעות: {ev['start_time']}-{ev['end_time']}")
                lines.append("")

                lines.append("ברמנים:")
                for name in ev.get("bartender_names") or ["—"]:
                    lines.append(f"  • {name}")

                lines.append("")
                lines.append("מלצרים:")
                for name in ev.get("waiter_names") or ["—"]:
                    lines.append(f"  • {name}")

                mb = ev.get("missing_bartenders", 0)
                mw = ev.get("missing_waiters", 0)
                if mb > 0 or mw > 0:
                    lines.append("")
                    lines.append(f"חסר: {mb} ברמנים, {mw} מלצרים")

                lines.append("")
                lines.append("—" * 18)
                lines.append("")

        content_cell = ws[f"{col}3"]
        content_cell.value = "\n".join(lines).strip()
        content_cell.font = normal_font
        content_cell.alignment = right_top
        content_cell.border = border

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ======================
# וואטסאפ: הודעה + HTML (Bytes)
# ======================
def make_personal_message(employee_name: str, shifts: list[dict]) -> str:
    LTR = "\u200e"
    days_order = ["א", "ב", "ג", "ד", "ה", "ו", "ש"]

    shifts_sorted = sorted(
        shifts,
        key=lambda x: (days_order.index(x["day"]), x["arrival_min"], x["start_min"])
    )

    if not shifts_sorted:
        return (
            f"{employee_name} היי,\n"
            "השבוע לא שובצת.\n"
            "אם יש שינוי אני אעדכן."
        )

    lines = [f"{employee_name} היי,", "זה הסידור שלך השבוע:", ""]

    current_day = None
    for ev in shifts_sorted:
        if ev["day"] != current_day:
            current_day = ev["day"]
            lines.append(f"יום {current_day}:")

        lines.append(f"הגעה: {LTR}{ev['arrival_time']}{LTR}")
        lines.append(f"אולם: {LTR}{ev['hall']}{LTR}")
        lines.append(f"שעות: {LTR}{ev['start_time']}-{ev['end_time']}{LTR}")
        lines.append("")

    lines.append("אם יש בעיה / איחור – תעדכן מראש.")
    return "\n".join(lines).strip()

def build_employee_shifts_map(week: list[dict], employees: list[dict]) -> dict:
    shifts = {e["name"]: [] for e in employees}
    for ev in week:
        for name in ev.get("waiter_names", []):
            if name in shifts:
                shifts[name].append(ev)
        for name in ev.get("bartender_names", []):
            if name in shifts:
                shifts[name].append(ev)
    return shifts

def build_whatsapp_html_bytes(employees: list[dict], week: list[dict]) -> bytes:
    shifts_map = build_employee_shifts_map(week, employees)

    cards_html = []
    for e in sorted(employees, key=lambda x: (x["role"], x["name"])):
        name = e["name"]
        phone = (e.get("phone") or "").strip()

        msg = make_personal_message(name, shifts_map.get(name, []))
        msg_encoded = quote(msg)

        ok_phone = phone.isdigit() and (8 <= len(phone) <= 15)
        wa_link = f"https://wa.me/{phone}?text={msg_encoded}" if ok_phone else ""

        role_he = "מלצר/ית" if e["role"] == "waiter" else "ברמן/ית"
        preview = (
            msg.replace("&", "&amp;")
               .replace("<", "&lt;")
               .replace(">", "&gt;")
        )

        if ok_phone:
            button = f'<a class="btn" href="{wa_link}" target="_blank" rel="noopener">שלח בוואטסאפ</a>'
        else:
            button = '<span class="bad">מספר לא תקין (תתקן ברשימת עובדים)</span>'

        cards_html.append(f"""
        <div class="card">
          <div class="top">
            <div class="name">{name}</div>
            <div class="meta">{role_he} • {phone if phone else "ללא מספר"}</div>
          </div>
          <div class="actions">{button}</div>
          <pre class="preview">{preview}</pre>
        </div>
        """)

    html = f"""<!doctype html>
<html lang="he" dir="rtl">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>שליחת סידור בוואטסאפ</title>
  <style>
    body {{ font-family: Arial, sans-serif; margin: 0; background: #f6f6f7; color: #111; }}
    header {{ position: sticky; top: 0; background: white; padding: 14px 16px; border-bottom: 1px solid #ddd; z-index: 10; }}
    h1 {{ margin: 0; font-size: 18px; }}
    .sub {{ margin-top: 6px; font-size: 13px; color: #444; line-height: 1.35; }}
    .wrap {{ padding: 12px; max-width: 900px; margin: 0 auto; }}
    .card {{ background: white; border: 1px solid #e2e2e2; border-radius: 12px; padding: 12px; margin-bottom: 12px; box-shadow: 0 1px 2px rgba(0,0,0,0.04); }}
    .top {{ display: flex; justify-content: space-between; gap: 10px; align-items: baseline; flex-wrap: wrap; }}
    .name {{ font-size: 18px; font-weight: 700; }}
    .meta {{ font-size: 13px; color: #555; }}
    .actions {{ margin-top: 10px; display: flex; gap: 10px; align-items: center; flex-wrap: wrap; }}
    .btn {{ display: inline-block; background: #25D366; color: white; padding: 10px 12px; border-radius: 10px; text-decoration: none; font-weight: 700; font-size: 14px; }}
    .bad {{ color: #b00020; font-weight: 700; font-size: 13px; }}
    .preview {{ margin-top: 10px; background: #fafafa; border: 1px dashed #ddd; border-radius: 10px; padding: 10px; white-space: pre-wrap; line-height: 1.35; font-size: 13px; color: #222; }}
    .note {{ margin-top: 10px; background: #fff7d6; border: 1px solid #f2d070; border-radius: 10px; padding: 10px; font-size: 13px; line-height: 1.35; }}
  </style>
</head>
<body>
  <header>
    <h1>שליחת סידור בוואטסאפ</h1>
    <div class="sub">
      לוחצים על הכפתור ליד עובד → וואטסאפ נפתח עם הודעה מוכנה → לוחצים "שלח".
    </div>
  </header>

  <div class="wrap">
    <div class="note">
      <b>חשוב:</b> מספרים חייבים להיות בפורמט בינלאומי בלי פלוס/רווחים/מקפים (למשל: 9725XXXXXXXX).
    </div>

    {''.join(cards_html)}
  </div>
</body>
</html>
"""
    return html.encode("utf-8")


# ======================
# Excel Import/Export (Employees)
# ======================
def employees_to_excel_bytes(employees: list[dict]) -> bytes:
    # נוח למנהל: days בתור "ב,ד,ה"
    rows = []
    for e in employees:
        rows.append({
            "name": e.get("name", ""),
            "phone": e.get("phone", ""),
            "role": e.get("role", ""),
            "days": ",".join(e.get("days", [])) if isinstance(e.get("days"), list) else str(e.get("days", "")),
            "rank": int(e.get("rank", 1)),
        })
    df = pd.DataFrame(rows, columns=["name", "phone", "role", "days", "rank"])
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="employees")
    return bio.getvalue()

def employees_from_excel(file) -> list[dict]:
    df = pd.read_excel(file)
    required = {"name", "phone", "role", "days", "rank"}
    if not required.issubset(set(df.columns)):
        missing = required - set(df.columns)
        raise ValueError(f"חסרות עמודות: {', '.join(missing)}")

    employees = []
    for _, r in df.iterrows():
        name = str(r["name"]).strip()
        phone = str(r["phone"]).strip()
        role = str(r["role"]).strip()
        days = [d.strip() for d in str(r["days"]).split(",") if d.strip()]
        rank = int(r["rank"])

        if not name:
            continue

        employees.append({
            "name": name,
            "phone": phone,
            "role": role,
            "days": days,
            "rank": rank,
        })
    return employees


# ======================
# ולידציה לשעות
# ======================
def is_time_ok(t: str) -> bool:
    t = (t or "").strip()
    if len(t) != 5 or t[2] != ":":
        return False
    hh, mm = t.split(":")
    if not (hh.isdigit() and mm.isdigit()):
        return False
    h = int(hh)
    m = int(mm)
    return 0 <= h <= 23 and 0 <= m <= 59


# ======================
# UI
# ======================
st.set_page_config(page_title="שיבוץ אירועים", layout="wide")
st.title("מערכת שיבוץ לאירועים")

init_db()
if not db_has_employees():
    seed_default_employees()

employees = load_employees_from_db()

# ---- ניהול עובדים ----
with st.expander("👥 ניהול עובדים (המנהל יכול לערוך כאן)", expanded=False):
    st.write("ערוך את הטבלה, הוסף/מחק שורות, ואז לחץ **שמור שינויים**. (נשמר קבוע ב־DB)")

    df = pd.DataFrame([{
        "name": e["name"],
        "phone": e["phone"],
        "role": e["role"],
        "days": ",".join(e["days"]),
        "rank": e["rank"],
    } for e in employees])

    edited = st.data_editor(
        df,
        use_container_width=True,
        num_rows="dynamic",
        column_config={
            "role": st.column_config.SelectboxColumn(
                "role",
                options=["waiter", "bartender"],
                required=True,
            ),
        }
    )

    cA, cB, cC = st.columns([1, 1, 2])

    with cA:
        if st.button("💾 שמור שינויים", type="primary"):
            new_emps = []
            errors = []

            for i, r in edited.iterrows():
                name = str(r.get("name", "")).strip()
                phone = str(r.get("phone", "")).strip()
                role = str(r.get("role", "")).strip()
                days = [d.strip() for d in str(r.get("days", "")).split(",") if d.strip()]
                rank_val = r.get("rank", 1)

                try:
                    rank = int(rank_val)
                except Exception:
                    rank = 1

                if not name:
                    continue
                if role not in ("waiter", "bartender"):
                    errors.append(f"שורה {i+1}: role חייב waiter/bartender")
                    continue

                new_emps.append({
                    "name": name,
                    "phone": phone,
                    "role": role,
                    "days": days,
                    "rank": rank,
                })

            if errors:
                st.error("יש בעיות:\n" + "\n".join(errors))
            else:
                save_employees_to_db(new_emps)
                st.success("נשמר! רענן/פתח מחדש ותראה שזה נשאר.")
                employees = load_employees_from_db()

    with cB:
        if st.button("♻️ החזר לברירת מחדל"):
            save_employees_to_db(build_default_employees())
            st.warning("הוחזר לברירת מחדל.")
            employees = load_employees_from_db()

    with cC:
        st.caption("טיפ: days בפורמט: ב,ד,ה,ו  | טלפון: 9725XXXXXXXX")

    st.divider()
    st.subheader("📦 ייבוא/ייצוא עובדים (Excel)")
    template_bytes = employees_to_excel_bytes(employees)
    st.download_button("⬇️ הורד אקסל עובדים", template_bytes, "employees.xlsx")

    upl = st.file_uploader("⬆️ העלה אקסל עובדים (xlsx)", type=["xlsx"])
    if upl:
        if st.button("ייבא והחלף את כל העובדים מהאקסל"):
            try:
                new_emps = employees_from_excel(upl)
                save_employees_to_db(new_emps)
                st.success("יובא ונשמר!")
                employees = load_employees_from_db()
            except Exception as ex:
                st.error(f"ייבוא נכשל: {ex}")

st.divider()

# ---- אירועים ----
if "events" not in st.session_state:
    st.session_state.events = []

days = ["א", "ב", "ג", "ד", "ה", "ו", "ש"]
halls = ["A", "B"]

st.subheader("➕ הוספת אירוע")
c1, c2, c3, c4, c5 = st.columns([1.1, 1.1, 1.2, 1.2, 1.4])

with c1:
    day = st.selectbox("יום", days)
with c2:
    hall = st.selectbox("אולם", halls)
with c3:
    guests = st.number_input("מוזמנים", min_value=1, max_value=3000, value=150, step=10)
with c4:
    start_time = st.text_input("שעת התחלה (HH:MM)", value="19:00")
with c5:
    end_time = st.text_input("שעת סיום (HH:MM)", value="01:00")

b1, b2, b3 = st.columns([1, 1, 2])
with b1:
    if st.button("הוסף אירוע"):
        if not is_time_ok(start_time) or not is_time_ok(end_time):
            st.error("שעה לא תקינה. דוגמה: 19:30")
        else:
            st.session_state.events.append({
                "day": day,
                "hall": hall,
                "guests": int(guests),
                "start_time": start_time.strip(),
                "end_time": end_time.strip(),
            })
            st.success("נוסף אירוע")
with b2:
    if st.button("נקה הכל"):
        st.session_state.events = []
        st.warning("נוקו כל האירועים")
with b3:
    st.caption("טיפ: תוסיף את כל האירועים ואז תלחץ 'צור סידור'.")

st.subheader("📋 רשימת אירועים שהוספת")
if not st.session_state.events:
    st.info("אין אירועים עדיין.")
else:
    st.dataframe(st.session_state.events, use_container_width=True)

st.divider()

# ---- יצירת סידור + הורדות ----
st.subheader("✅ יצירת סידור והורדת קבצים")
if st.button("צור סידור עכשיו", type="primary"):
    if not st.session_state.events:
        st.error("אין אירועים ליצור מהם סידור.")
    else:
        week = [
            build_event(e["day"], e["hall"], e["guests"], e["start_time"], e["end_time"])
            for e in st.session_state.events
        ]

        assigned = {}
        schedule = {}
        for ev in week:
            assign_event(ev, employees, assigned, schedule)

        by_day = {}
        for ev in week:
            by_day.setdefault(ev["day"], []).append(ev)

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        xlsx_name = f"סידור_שבועי_למנהל_{stamp}.xlsx"
        html_name = f"שליחת_סידור_וואטסאפ_{stamp}.html"

        xlsx_bytes = build_manager_excel_bytes(by_day)
        html_bytes = build_whatsapp_html_bytes(employees, week)

        st.success("נוצר! תוריד את הקבצים 👇")

        d1, d2 = st.columns(2)
        with d1:
            st.download_button(
                label="⬇️ הורד אקסל למנהל",
                data=xlsx_bytes,
                file_name=xlsx_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        with d2:
            st.download_button(
                label="⬇️ הורד דף כפתורי וואטסאפ (HTML)",
                data=html_bytes,
                file_name=html_name,
                mime="text/html",
            )

        st.info(
            "איך משתמשים ב-HTML:\n"
            "1) פותחים את הקובץ בדפדפן (מחשב או טלפון)\n"
            "2) לוחצים על 'שלח בוואטסאפ' ליד העובד\n"
            "3) וואטסאפ נפתח עם הודעה מוכנה וברורה\n"
            "4) לוחצים 'שלח'"
        )